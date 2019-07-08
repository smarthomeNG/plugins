#!/usr/bin/env python3
# vim: set encoding=utf-8 tabstop=4 softtabstop=4 shiftwidth=4 expandtab
#########################################################################
#  Copyright 2016 - 2018 Bernd Meiners              Bernd.Meiners@mail.de
#########################################################################
#
#  DLMS plugin for SmartHomeNG.py.
#
#  This file is part of SmartHomeNG.py.
#  Visit:  https://github.com/smarthomeNG/
#          https://knx-user-forum.de/forum/supportforen/smarthome-py
#          https://smarthomeng.de
#
#  SmartHomeNG.py is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  SmartHomeNG.py is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with SmartHomeNG.py. If not, see <http://www.gnu.org/licenses/>.
#########################################################################

"""
On standalone mode this Python program will visit
https://www.dlms.com/srv/lib/Export_Flagids.php
to download the list of manufacturer ids for smartmeter as xlsx
The columns therein are labeled as "FLAG ID","Manufacturer", "Country", and "World Region"
``FLAG ID`` is exactly 3 characters long

The result will be stored locally as ``manufacturer.yaml``
to serve as information database for the identification of smartmeters

"""

import logging
import requests

from ruamel.yaml import YAML
from io import BytesIO
import openpyxl

if __name__ == '__main__':
    logger = logging.getLogger(__name__)
    logger.debug("init standalone {}".format(__name__))
else:
    logger = logging.getLogger()
    logger.debug("init plugin component {}".format(__name__))


def get_manufacturer( from_url, to_yaml, verbose = False ):
    """
    Read XLSX from given url and write a yaml containing id and manufacturer
    """
    # result
    r = {}
    y = YAML()

    logger.debug("Read manufacturer IDs from URL: '{}'".format(url))
    headers = {'User-agent': 'Mozilla/5.0'}

    try:
        reque = requests.get(url, headers=headers)
    except ConnectionError as e:
        logger.debug('An error occurred fetching {} \n {}'.format(url, e.reason))
        raise

    try:
        wb = openpyxl.load_workbook(filename=BytesIO(reque.content), data_only=True)
        #wb = openpyxl.load_workbook(xlfilename, data_only=True)

        logger.debug('sheetnames {}'.format(wb.get_sheet_names()))
        
        sheet = wb.active
        logger.debug('sheet {}'.format(sheet))
        logger.debug('rows [{} ..{}]'.format(sheet.min_row, sheet.max_row))
        logger.debug('columns [{} ..{}]'.format(sheet.min_column, sheet.max_column))
        
        if sheet.min_row+1 <= sheet.max_row and sheet.min_column == 1 and sheet.max_column == 4:
            # Get data from rows """
            for row in range(sheet.min_row+1,sheet.max_row):
                id = sheet.cell(row, 1).value
                man = sheet.cell(row, 2).value
                r[id] = man
                if verbose:
                    logger.debug("{}->{}".format(id,man))
            with open(exportfile, 'w') as f:
                y.dump( r, f )

        logger.debug("{} distinct manufacturers were found and written to {}".format(len(r),exportfile))
        
    except Exception as e:
        logger.debug("Error {} occurred".format(e))

    return r

if __name__ == '__main__':
    verbose = False

    logging.getLogger().setLevel( logging.DEBUG )
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    # create formatter and add it to the handlers
    if verbose:
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s  @ %(lineno)d')
    else:
        formatter = logging.Formatter('%(message)s')
    ch.setFormatter(formatter)
    # add the handlers to the logger
    logging.getLogger().addHandler(ch)
    logger = logging.getLogger(__name__)

    exportfile = 'manufacturer.yaml'
    url = 'https://www.dlms.com/srv/lib/Export_Flagids.php'
    get_manufacturer( url, exportfile, verbose)

