#!/usr/bin/env python3
# vim: set encoding=utf-8 tabstop=4 softtabstop=4 shiftwidth=4 expandtab
#########################################################################
#  Copyright 2016 - 2021 Bernd Meiners              Bernd.Meiners@mail.de
#########################################################################
#
#  DLMS plugin for SmartHomeNG
#
#  This file is part of SmartHomeNG.py.
#  Visit:  https://github.com/smarthomeNG/
#          https://knx-user-forum.de/forum/supportforen/smarthome-py
#          https://smarthomeng.de
#
#  SmartHomeNG.py is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  SmartHomeNG.py is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with SmartHomeNG.py. If not, see <http://www.gnu.org/licenses/>.
#########################################################################

"""
On standalone mode this Python program will visit
https://www.dlms.com/srv/lib/Export_Flagids.php
to download the list of manufacturer ids for smartmeter as xlsx
The columns therein are labeled as "FLAG ID","Manufacturer", "Country", and "World Region"
``FLAG ID`` is exactly 3 characters long

The result will be stored locally as ``manufacturer.yaml``
to serve as information database for the identification of smartmeters

"""

import logging
import requests
import sys

from ruamel.yaml import YAML
from io import BytesIO
install_openpyxl = "python3 -m pip install --user openpyxl"

if __name__ == '__main__':
    logger = logging.getLogger(__name__)
    logger.debug(f"init standalone {__name__}")
else:
    logger = logging.getLogger()
    logger.debug(f"init plugin component {__name__}")

try:
    import openpyxl
except:
    sys.exit(f"Package 'openpyxl' was not found. You might install with {install_openpyxl}")


def get_manufacturer( from_url, to_yaml, verbose = False ):
    """
    Read XLSX from given url and write a yaml containing id and manufacturer
    """
    # result
    r = {}
    y = YAML()

    logger.debug(f"Read manufacturer IDs from URL: '{url}'")
    logger.debug(f"Using openpyxl version '{openpyxl.__version__}'")
    
    headers = {'User-agent': 'Mozilla/5.0'}

    try:
        reque = requests.get(url, headers=headers)
    except ConnectionError as e:
        logger.debug(f"An error {e} occurred fetching {url}\n")
        raise

    try:
        wb = openpyxl.load_workbook(filename=BytesIO(reque.content), data_only=True)
        #wb = openpyxl.load_workbook(xlfilename, data_only=True)

        logger.debug('sheetnames {}'.format(wb.sheetnames))
        
        sheet = wb.active
        logger.debug(f"sheet {sheet}")
        logger.debug(f"rows [{sheet.min_row} .. {sheet.max_row}]")
        logger.debug(f"columns [{sheet.min_column} .. {sheet.max_column}]")
        
        if sheet.min_row+1 <= sheet.max_row and sheet.min_column == 1 and sheet.max_column == 4:
            # Get data from rows """
            for row in range(sheet.min_row+1,sheet.max_row):
                id = str(sheet.cell(row, 1).value).strip()
                if len(id) == 3:
                    # there are entries like > 'ITRON ...'  < that need special cleaning:
                    man = str(sheet.cell(row, 2).value).strip()
                    man = man.strip('\'').strip()
                    r[id] = man
                    if verbose:
                        logger.debug(f"{id}->{man}")
                else:
                    logger.debug(f">id< is '{id}' has more than 3 characters and will not be considered")
            with open(exportfile, 'w') as f:
                y.dump( r, f )

        logger.debug(f"{len(r)} distinct manufacturers were found and written to {exportfile}")
        
    except Exception as e:
        logger.debug(f"Error {e} occurred")

    return r

if __name__ == '__main__':
    verbose = True

    logging.getLogger().setLevel( logging.DEBUG )
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    # create formatter and add it to the handlers
    if verbose:
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s  @ %(lineno)d')
    else:
        formatter = logging.Formatter('%(message)s')
    ch.setFormatter(formatter)
    # add the handlers to the logger
    logging.getLogger().addHandler(ch)
    logger = logging.getLogger(__name__)

    exportfile = 'manufacturer.yaml'
    url = 'https://www.dlms.com/srv/lib/Export_Flagids.php'
    get_manufacturer( url, exportfile, verbose)

